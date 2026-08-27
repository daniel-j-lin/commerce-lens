"""Read-only `.xlsx` inspection adapter."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from commerce_lens.contracts.common import FailureDetail, FailureStage, SourceType
from commerce_lens.intake.inspection import ColumnInspection, InspectionStatus, IntakeInspectionResult, infer_observed_type
from commerce_lens.intake.registry import DatasetRegistry


class ExcelInspectionAdapter:
    def __init__(self, registry: DatasetRegistry | None = None, *, sample_rows: int = 50) -> None:
        self.registry = registry
        self.sample_rows = sample_rows

    def inspect(self, source_path: str | Path, *, sheet_name: str | None = None) -> IntakeInspectionResult:
        path = Path(source_path)
        if path.suffix.lower() == ".xls":
            return _failure("legacy .xls files are unsupported in Phase 1", InspectionStatus.UNSUPPORTED)
        if path.suffix.lower() != ".xlsx":
            return _failure("only .xlsx Excel workbooks are supported in Phase 1", InspectionStatus.UNSUPPORTED)
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            formula_workbook = load_workbook(path, read_only=False, data_only=False)
        except Exception as exc:
            return _failure(f"Excel workbook is unreadable: {exc}", InspectionStatus.FAILED)

        sheets = tuple(workbook.sheetnames)
        if sheet_name is None:
            if len(sheets) != 1:
                return IntakeInspectionResult(
                    source_type=SourceType.EXCEL_XLSX,
                    status=InspectionStatus.AMBIGUOUS,
                    available_sheets=sheets,
                    ambiguities=("multiple sheets require explicit sheet selection",),
                    failure_detail=FailureDetail(stage=FailureStage.INTAKE, reason="multiple sheets require explicit sheet selection"),
                )
            sheet_name = sheets[0]
        if sheet_name not in sheets:
            return _failure(f"selected sheet does not exist: {sheet_name}", InspectionStatus.FAILED, sheets=sheets)

        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True, max_row=self.sample_rows + 1))
        if not rows:
            return _failure("selected sheet contains no rows", InspectionStatus.FAILED, sheets=sheets)
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        header_failure = _validate_headers(headers)
        if header_failure:
            return _failure(header_failure, InspectionStatus.FAILED, sheets=sheets)
        formula_failure = _missing_formula_stored_value(
            formula_workbook[sheet_name],
            rows,
            max_columns=len(headers),
        )
        if formula_failure is not None:
            return _failure(formula_failure, InspectionStatus.FAILED, sheets=sheets)
        data_rows = [tuple(row) for row in rows[1:]]
        columns: list[ColumnInspection] = []
        for index, name in enumerate(headers):
            values = [row[index] if index < len(row) else None for row in data_rows]
            observed_type, nullable = infer_observed_type(values)
            columns.append(ColumnInspection(name=name, position=index, observed_type=observed_type, nullable_observed=nullable))

        warnings = _formula_warnings(
            formula_workbook[sheet_name],
            rows_to_check=len(rows),
            max_columns=len(headers),
        )
        dataset_id = None
        if self.registry is not None:
            dataset_id = self.registry.register_source(path, SourceType.EXCEL_XLSX, selected_sheet=sheet_name).dataset_id
        return IntakeInspectionResult(
            source_type=SourceType.EXCEL_XLSX,
            status=InspectionStatus.SUPPORTED,
            dataset_ref_id=dataset_id,
            selected_sheet=sheet_name,
            available_sheets=sheets,
            columns=tuple(columns),
            row_count=max(0, sheet.max_row - 1),
            warnings=warnings,
        )


def _missing_formula_stored_value(sheet, data_only_rows: list[tuple[object, ...]], *, max_columns: int) -> str | None:
    for row_index, row in enumerate(sheet.iter_rows(max_row=len(data_only_rows), max_col=max_columns), start=1):
        for column_index, cell in enumerate(row, start=1):
            if cell.data_type != "f":
                continue
            data_only_value = None
            if row_index <= len(data_only_rows) and column_index <= len(data_only_rows[row_index - 1]):
                data_only_value = data_only_rows[row_index - 1][column_index - 1]
            if data_only_value is None:
                return (
                    f"formula cell {cell.coordinate} has no stored value; "
                    "CommerceLens does not recalculate Excel formulas"
                )
    return None


def _formula_warnings(sheet, *, rows_to_check: int, max_columns: int) -> tuple[str, ...]:
    formula_cells = []
    for row in sheet.iter_rows(max_row=rows_to_check, max_col=max_columns):
        for cell in row:
            if cell.data_type == "f":
                formula_cells.append(cell.coordinate)
    if not formula_cells:
        return ()
    return ("formula cells are inspected from stored values only; workbook formulas are not recalculated",)


def _validate_headers(headers: list[str]) -> str | None:
    if not headers or any(not header for header in headers):
        return "Excel header row contains empty column names"
    if len(set(headers)) != len(headers):
        return "Excel header row contains duplicate column names"
    return None


def _failure(reason: str, status: InspectionStatus, *, sheets: tuple[str, ...] = ()) -> IntakeInspectionResult:
    return IntakeInspectionResult(
        source_type=SourceType.EXCEL_XLSX,
        status=status,
        available_sheets=sheets,
        failure_detail=FailureDetail(stage=FailureStage.INTAKE, reason=reason),
        ambiguities=(reason,) if status is InspectionStatus.AMBIGUOUS else (),
    )
