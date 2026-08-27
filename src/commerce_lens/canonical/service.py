"""Deterministic canonicalization service for Phase 2."""

from __future__ import annotations

import csv
import math
import sqlite3
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import load_workbook

from commerce_lens.canonical.mapping import validate_mapping
from commerce_lens.canonical.models import (
    CANONICALIZATION_VERSION,
    SUPPORTED_DATE_POLICY_REF,
    CanonicalLineRecord,
    CanonicalizationRequest,
    CanonicalizationResult,
    EligibilityMode,
    EligibilityState,
)
from commerce_lens.canonical.quality import (
    DataQualityCheckResult,
    DataQualityConsequence,
    blocking,
    passed,
    qualification,
)
from commerce_lens.canonical.schema import (
    CANONICAL_FIELD_NAMES,
    CANONICAL_SCHEMA_VERSION,
    UNCLASSIFIED_CATEGORY_ID,
    UNCLASSIFIED_CATEGORY_NAME,
)
from commerce_lens.contracts.common import ArtifactReference, Qualification
from commerce_lens.contracts.evidence import CanonicalDatasetReference, CanonicalizationRecord, DatasetReference
from commerce_lens.evidence.identifiers import canonical_json_fingerprint, sha256_file, stable_content_id
from commerce_lens.persistence.artifact_store import ArtifactStore


def canonicalize_dataset(
    dataset: DatasetReference,
    request: CanonicalizationRequest,
    artifact_store: ArtifactStore,
) -> CanonicalizationResult:
    if dataset.dataset_id != request.source_dataset_id:
        raise ValueError("canonicalization request source_dataset_id does not match dataset reference")
    source_path = _dataset_source_path(dataset, artifact_store)
    quality_results: list[DataQualityCheckResult] = list(_validate_request_configuration(dataset, request, source_path))
    if any(result.consequence is DataQualityConsequence.BLOCKING for result in quality_results):
        return _blocked_result(dataset, request, quality_results)
    selected_sheet = _governed_selected_sheet(dataset, source_path)
    selected_table = _governed_selected_table(dataset, source_path)

    source_rows = _read_source_rows(
        dataset,
        source_path,
        selected_sheet=selected_sheet,
        selected_table=selected_table,
        order_date_source_field=request.mapping.source_for("order_date"),
    )
    source_columns = tuple(source_rows[0].keys()) if source_rows else _source_columns(
        dataset,
        source_path,
        selected_sheet=selected_sheet,
        selected_table=selected_table,
    )

    quality_results.extend(
        validate_mapping(
            request.mapping,
            source_columns,
            require_category=request.require_category,
            require_eligibility=request.eligibility_mode is EligibilityMode.EXPLICIT_STATUS_MAPPING,
        )
    )
    quality_results.extend(_validate_eligibility_configuration(request))
    if request.unsupported_partial_refund_evidence:
        quality_results.append(
            blocking(
                "canonical.partial_refund_unsupported",
                dataset.dataset_id,
                "canonical_dictionary:12",
                "explicit partial-refund evidence exists without a governed final line_revenue",
            )
        )
    if any(result.consequence is DataQualityConsequence.BLOCKING for result in quality_results):
        return _blocked_result(dataset, request, quality_results)

    canonical_rows: list[CanonicalLineRecord] = []
    for index, source_row in enumerate(source_rows, start=2):
        parsed, row_results = _canonicalize_row(source_row, index, request)
        quality_results.extend(row_results)
        if parsed is not None:
            canonical_rows.append(parsed)

    quality_results.extend(_cross_row_quality(canonical_rows, request))
    qualifications = _qualifications_from_quality(quality_results)
    blocking_results = [result for result in quality_results if result.consequence is DataQualityConsequence.BLOCKING]
    if blocking_results:
        return _blocked_result(dataset, request, quality_results, qualifications=qualifications)

    quality_results.extend(_monetary_artifact_quality(canonical_rows))
    if any(result.consequence is DataQualityConsequence.BLOCKING for result in quality_results):
        return _blocked_result(dataset, request, quality_results, qualifications=qualifications)
    quality_results.append(
        passed(
            "canonical.dataset.valid",
            dataset.dataset_id,
            "phase2:canonicalization",
            "canonical dataset satisfies Phase 2 field and cross-row checks",
        )
    )
    canonical_dataset = _write_canonical_artifact(
        dataset,
        request,
        canonical_rows,
        artifact_store,
        selected_sheet=selected_sheet,
        selected_table=selected_table,
    )
    record = _record(
        dataset=dataset,
        request=request,
        canonical_dataset_id=canonical_dataset.canonical_dataset_id,
        output_fingerprint=canonical_dataset.content_fingerprint,
        source_row_count=len(source_rows),
        canonical_row_count=len(canonical_rows),
        quality_results=quality_results,
        qualifications=qualifications,
        failures=(),
        selected_sheet=selected_sheet,
        selected_table=selected_table,
    )
    return CanonicalizationResult(
        record=record,
        canonical_dataset=canonical_dataset,
        data_quality_results=tuple(quality_results),
        qualifications=qualifications,
        canonical_rows=tuple(canonical_rows),
    )


def _dataset_source_path(dataset: DatasetReference, artifact_store: ArtifactStore) -> Path:
    if dataset.snapshot_artifact is None:
        raise ValueError("canonicalization requires an immutable source snapshot")
    source_path = artifact_store.safe_path(dataset.snapshot_artifact.path)
    if sha256_file(source_path) != dataset.content_fingerprint:
        raise ValueError("source snapshot fingerprint does not match dataset reference")
    return source_path


def _validate_request_configuration(
    dataset: DatasetReference,
    request: CanonicalizationRequest,
    source_path: Path,
) -> tuple[DataQualityCheckResult, ...]:
    results: list[DataQualityCheckResult] = []
    if request.canonical_schema_version != CANONICAL_SCHEMA_VERSION:
        results.append(
            blocking(
                "canonical.config.unsupported_schema",
                request.canonical_schema_version,
                "canonical_dictionary:9",
                "unsupported canonical schema version",
            )
        )
    if request.date_policy_ref != SUPPORTED_DATE_POLICY_REF:
        results.append(
            blocking(
                "canonical.config.unsupported_date_policy",
                request.date_policy_ref,
                "canonical_dictionary:16",
                "unsupported canonical date policy",
            )
        )
    if dataset.selected_sheet is not None and request.selected_sheet is not None and dataset.selected_sheet != request.selected_sheet:
        results.append(
            blocking(
                "canonical.selection.conflicting_sheet",
                request.selected_sheet,
                "architecture:9.1",
                "CanonicalizationRequest selected_sheet conflicts with DatasetReference selected_sheet",
            )
        )
    if dataset.selected_table is not None and request.selected_table is not None and dataset.selected_table != request.selected_table:
        results.append(
            blocking(
                "canonical.selection.conflicting_table",
                request.selected_table,
                "architecture:9.1",
                "CanonicalizationRequest selected_table conflicts with DatasetReference selected_table",
            )
        )
    if dataset.selected_sheet is None and request.selected_sheet is not None:
        results.append(
            blocking(
                "canonical.selection.request_sheet_without_dataset_selection",
                request.selected_sheet,
                "architecture:9.1",
                "CanonicalizationRequest must not supply a sheet selection absent from DatasetReference; re-register the source with governed sheet selection",
            )
        )
    if dataset.selected_table is None and request.selected_table is not None:
        results.append(
            blocking(
                "canonical.selection.request_table_without_dataset_selection",
                request.selected_table,
                "architecture:9.1",
                "CanonicalizationRequest must not supply a table selection absent from DatasetReference; re-register the source with governed table selection",
            )
        )
    if dataset.source_type.value == "excel_xlsx" and dataset.selected_sheet is None and request.selected_sheet is None:
        sheet_names = _excel_sheet_names(source_path)
        if len(sheet_names) != 1:
            results.append(
                blocking(
                    "canonical.selection.missing_governed_sheet",
                    dataset.dataset_id,
                    "architecture:9.1",
                    "multi-sheet workbook requires governed sheet selection in DatasetReference",
                )
            )
    if dataset.source_type.value == "sqlite" and dataset.selected_table is None and request.selected_table is None:
        table_names = _sqlite_table_names(source_path)
        if len(table_names) != 1:
            results.append(
                blocking(
                    "canonical.selection.missing_governed_table",
                    dataset.dataset_id,
                    "architecture:9.1",
                    "multi-table SQLite source requires governed table selection in DatasetReference",
                )
            )
    return tuple(results)


def _effective_selection(dataset_selection: str | None, request_selection: str | None) -> str | None:
    return dataset_selection if dataset_selection is not None else request_selection


def _governed_selected_sheet(dataset: DatasetReference, source_path: Path) -> str | None:
    if dataset.source_type.value != "excel_xlsx":
        return None
    if dataset.selected_sheet is not None:
        return dataset.selected_sheet
    sheet_names = _excel_sheet_names(source_path)
    return sheet_names[0] if len(sheet_names) == 1 else None


def _governed_selected_table(dataset: DatasetReference, source_path: Path) -> str | None:
    if dataset.source_type.value != "sqlite":
        return None
    if dataset.selected_table is not None:
        return dataset.selected_table
    table_names = _sqlite_table_names(source_path)
    return table_names[0] if len(table_names) == 1 else None


def _excel_sheet_names(source_path: Path) -> tuple[str, ...]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    return tuple(workbook.sheetnames)


def _sqlite_table_names(source_path: Path) -> tuple[str, ...]:
    conn = sqlite3.connect(f"{source_path.as_uri()}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    with conn:
        rows = conn.execute(
            """
            SELECT name
            FROM sqlite_master
            WHERE type IN ('table', 'view')
              AND name NOT LIKE 'sqlite_%'
            ORDER BY name
            """
        ).fetchall()
        return tuple(row["name"] for row in rows)


def _read_source_rows(
    dataset: DatasetReference,
    source_path: Path,
    *,
    selected_sheet: str | None,
    selected_table: str | None,
    order_date_source_field: str | None,
) -> list[dict[str, Any]]:
    if dataset.source_type.value == "csv":
        text = source_path.read_text(encoding="utf-8-sig")
        reader = csv.DictReader(text.splitlines())
        return [dict(row) for row in reader]
    if dataset.source_type.value == "excel_xlsx":
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        if selected_sheet is None:
            raise ValueError("canonicalization requires governed sheet selection")
        sheet = workbook[selected_sheet]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        return [
            {
                headers[index]: _normalize_excel_value(
                    headers[index],
                    value,
                    order_date_source_field=order_date_source_field,
                )
                for index, value in enumerate(row)
                if index < len(headers)
            }
            for row in rows[1:]
        ]
    if dataset.source_type.value == "sqlite":
        if selected_table is None:
            raise ValueError("canonicalization requires governed table selection")
        conn = sqlite3.connect(f"{source_path.as_uri()}?mode=ro", uri=True)
        conn.row_factory = sqlite3.Row
        with conn:
            quoted = '"' + selected_table.replace('"', '""') + '"'
            rows = conn.execute(f"SELECT * FROM {quoted}").fetchall()
            return [dict(row) for row in rows]
    raise ValueError(f"unsupported source type for canonicalization: {dataset.source_type.value}")


def _normalize_excel_value(source_field: str, value: Any, *, order_date_source_field: str | None) -> Any:
    if (
        source_field == order_date_source_field
        and isinstance(value, datetime)
        and value.time() == datetime.min.time()
    ):
        return value.date()
    return value


def _source_columns(
    dataset: DatasetReference,
    source_path: Path,
    *,
    selected_sheet: str | None,
    selected_table: str | None,
) -> tuple[str, ...]:
    if dataset.source_type.value == "csv":
        with source_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            return tuple(csv.DictReader(file_obj).fieldnames or ())
    if dataset.source_type.value == "excel_xlsx":
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        if selected_sheet is None:
            return ()
        first_row = next(workbook[selected_sheet].iter_rows(values_only=True, max_row=1), ())
        return tuple(str(value).strip() if value is not None else "" for value in first_row)
    if dataset.source_type.value == "sqlite":
        if selected_table is None:
            return ()
        conn = sqlite3.connect(f"{source_path.as_uri()}?mode=ro", uri=True)
        conn.row_factory = sqlite3.Row
        with conn:
            rows = conn.execute(f"PRAGMA table_info('{selected_table.replace(chr(39), chr(39) + chr(39))}')").fetchall()
            return tuple(row["name"] for row in rows)
    return ()


def _validate_eligibility_configuration(request: CanonicalizationRequest) -> tuple[DataQualityCheckResult, ...]:
    if request.eligibility_mode is EligibilityMode.UPSTREAM_ELIGIBLE_ONLY:
        return (
            passed(
                "canonical.eligibility.mode",
                request.source_dataset_id,
                "canonical_dictionary:12",
                "source is explicitly declared upstream-governed eligible-only",
            ),
        )
    states = {entry.normalized_status for entry in request.eligibility_value_mapping}
    if states != {EligibilityState.ELIGIBLE, EligibilityState.EXCLUDED}:
        return (
            blocking(
                "canonical.eligibility.mapping_incomplete",
                request.source_dataset_id,
                "canonical_dictionary:12",
                "explicit eligibility mapping must distinguish Eligible and Excluded",
            ),
        )
    source_values = [entry.source_value for entry in request.eligibility_value_mapping]
    if len(set(source_values)) != len(source_values):
        return (
            blocking(
                "canonical.eligibility.duplicate_source_value",
                request.source_dataset_id,
                "canonical_dictionary:12",
                "eligibility mapping contains a duplicate source status value",
            ),
        )
    return (
        passed(
            "canonical.eligibility.mapping_valid",
            request.source_dataset_id,
            "canonical_dictionary:12",
            "explicit eligibility mapping is governed and complete",
        ),
    )


def _canonicalize_row(
    source_row: dict[str, Any],
    source_row_number: int,
    request: CanonicalizationRequest,
) -> tuple[CanonicalLineRecord | None, tuple[DataQualityCheckResult, ...]]:
    results: list[DataQualityCheckResult] = []
    values = {field: source_row.get(request.mapping.source_for(field) or "") for field in CANONICAL_FIELD_NAMES}

    order_id = _required_identifier(values["order_id"], "order_id", source_row_number, results)
    order_line_id = _required_identifier(values["order_line_id"], "order_line_id", source_row_number, results)
    order_date = _parse_order_date(values["order_date"], source_row_number, results)
    product_id = _required_identifier(values["product_id"], "product_id", source_row_number, results)
    quantity = _parse_quantity(values["quantity"], source_row_number, results)
    line_revenue = _parse_decimal(values["line_revenue"], "line_revenue", source_row_number, results)
    currency = _required_identifier(values["currency"], "currency", source_row_number, results)
    unit_price = _parse_optional_decimal(values["unit_price"], "unit_price", source_row_number, results)
    eligibility = _eligibility_status(values["eligibility_status"], source_row_number, request, results)
    product_name = _optional_text(values["product_name"])
    category_id = _category_id(values["category_id"], source_row_number, request, results)
    category_name = _optional_text(values["category_name"])
    if request.require_category and category_id == UNCLASSIFIED_CATEGORY_ID:
        category_name = UNCLASSIFIED_CATEGORY_NAME
        results.append(
            qualification(
                "canonical.category.unclassified",
                f"row:{source_row_number}:category_id",
                "canonical_dictionary:20",
                "missing category_id is assigned to the governed Unclassified bucket",
            )
        )

    if any(result.consequence is DataQualityConsequence.BLOCKING for result in results):
        return None, tuple(results)
    assert order_id is not None
    assert order_line_id is not None
    assert order_date is not None
    assert product_id is not None
    assert quantity is not None
    assert line_revenue is not None
    assert currency is not None
    assert eligibility is not None
    return (
        CanonicalLineRecord(
            source_row_number=source_row_number,
            order_id=order_id,
            order_line_id=order_line_id,
            order_date=order_date,
            product_id=product_id,
            product_name=product_name,
            category_id=category_id,
            category_name=category_name,
            quantity=quantity,
            line_revenue=line_revenue,
            currency=currency,
            eligibility_status=eligibility,
            unit_price=unit_price,
        ),
        tuple(results),
    )


def _required_identifier(
    value: Any,
    field_name: str,
    row_number: int,
    results: list[DataQualityCheckResult],
) -> str | None:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        results.append(
            blocking(
                f"canonical.{field_name}.missing",
                f"row:{row_number}:{field_name}",
                "canonical_dictionary:9",
                f"{field_name} is required and must not be null",
            )
        )
        return None
    if isinstance(value, bool):
        results.append(
            blocking(
                f"canonical.{field_name}.invalid",
                f"row:{row_number}:{field_name}",
                "canonical_dictionary:9",
                f"{field_name} must be a stable non-boolean identifier",
            )
        )
        return None
    return str(value).strip()


def _parse_order_date(value: Any, row_number: int, results: list[DataQualityCheckResult]) -> date | None:
    if isinstance(value, datetime):
        results.append(
            blocking(
                "canonical.order_date.timestamp_unsupported",
                f"row:{row_number}:order_date",
                "canonical_dictionary:16",
                "timestamp values require upstream governed conversion to order_date",
            )
        )
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if len(text) == 10 and text[4] == "-" and text[7] == "-":
            try:
                return date.fromisoformat(text)
            except ValueError:
                pass
        results.append(
            blocking(
                "canonical.order_date.invalid",
                f"row:{row_number}:order_date",
                "canonical_dictionary:16",
                "order_date must be a native date or unambiguous ISO date string",
            )
        )
        return None
    results.append(
        blocking(
            "canonical.order_date.invalid",
            f"row:{row_number}:order_date",
            "canonical_dictionary:16",
            "order_date must be a native date or unambiguous ISO date string",
        )
    )
    return None


def _parse_quantity(value: Any, row_number: int, results: list[DataQualityCheckResult]) -> int | None:
    if value is None or value == "" or isinstance(value, bool):
        results.append(_quantity_failure(row_number, "quantity must be present as a positive whole number"))
        return None
    decimal = _decimal_from_source(value)
    if decimal is None:
        results.append(_quantity_failure(row_number, "quantity must be finite numeric input with integer semantics"))
        return None
    if decimal != decimal.to_integral_value() or decimal <= 0:
        results.append(_quantity_failure(row_number, "quantity must be positive and whole-number"))
        return None
    return int(decimal)


def _quantity_failure(row_number: int, reason: str) -> DataQualityCheckResult:
    return blocking(
        "canonical.quantity.invalid",
        f"row:{row_number}:quantity",
        "canonical_dictionary:31",
        reason,
    )


def _parse_decimal(
    value: Any,
    field_name: str,
    row_number: int,
    results: list[DataQualityCheckResult],
) -> Decimal | None:
    if value is None or value == "" or isinstance(value, bool):
        results.append(
            blocking(
                f"canonical.{field_name}.invalid",
                f"row:{row_number}:{field_name}",
                "canonical_dictionary:32.1",
                f"{field_name} must be present as an exact decimal value",
            )
        )
        return None
    decimal = _decimal_from_source(value)
    if decimal is None:
        results.append(
            blocking(
                f"canonical.{field_name}.invalid",
                f"row:{row_number}:{field_name}",
                "canonical_dictionary:32.1",
                f"{field_name} must be a valid finite Decimal-compatible value",
            )
        )
        return None
    if decimal < 0:
        results.append(
            blocking(
                f"canonical.{field_name}.negative",
                f"row:{row_number}:{field_name}",
                "canonical_dictionary:32.1",
                f"{field_name} must be non-negative",
            )
        )
        return None
    return decimal


def _decimal_from_source(value: Any) -> Decimal | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        text = str(value)
    else:
        text = str(value).strip()
    try:
        decimal = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    if not decimal.is_finite():
        return None
    return decimal


def _parse_optional_decimal(
    value: Any,
    field_name: str,
    row_number: int,
    results: list[DataQualityCheckResult],
) -> Decimal | None:
    if value is None or value == "":
        return None
    return _parse_decimal(value, field_name, row_number, results)


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _category_id(
    value: Any,
    row_number: int,
    request: CanonicalizationRequest,
    results: list[DataQualityCheckResult],
) -> str | None:
    if not request.require_category:
        return _optional_text(value)
    if isinstance(value, (list, tuple, set)):
        results.append(
            blocking(
                "canonical.category.multiple_assignment",
                f"row:{row_number}:category_id",
                "canonical_dictionary:20",
                "multi-category allocation is outside the MVP",
            )
        )
        return None
    text = _optional_text(value)
    return text or UNCLASSIFIED_CATEGORY_ID


def _eligibility_status(
    value: Any,
    row_number: int,
    request: CanonicalizationRequest,
    results: list[DataQualityCheckResult],
) -> EligibilityState | None:
    if request.eligibility_mode is EligibilityMode.UPSTREAM_ELIGIBLE_ONLY:
        return EligibilityState.ELIGIBLE
    if value is None or (isinstance(value, str) and value.strip() == ""):
        results.append(
            blocking(
                "canonical.eligibility.missing",
                f"row:{row_number}:eligibility_status",
                "canonical_dictionary:12",
                "eligibility status is required when the source contains eligible and excluded rows",
            )
        )
        return None
    text = str(value).strip()
    for entry in request.eligibility_value_mapping:
        if entry.source_value == text:
            return entry.normalized_status
    results.append(
        blocking(
            "canonical.eligibility.unknown",
            f"row:{row_number}:eligibility_status",
            "canonical_dictionary:12",
            "source eligibility value is not explicitly mapped",
        )
    )
    return None


def _cross_row_quality(
    rows: list[CanonicalLineRecord],
    request: CanonicalizationRequest,
) -> tuple[DataQualityCheckResult, ...]:
    results: list[DataQualityCheckResult] = []
    identities: dict[tuple[str, str], int] = {}
    duplicate_identity = False
    order_dates: dict[str, date] = {}
    inconsistent_order_dates = False
    product_names: dict[str, set[str]] = {}
    currencies: set[str] = set()
    for row in rows:
        identity = (row.order_id, row.order_line_id)
        if identity in identities:
            duplicate_identity = True
            results.append(
                blocking(
                    "canonical.identity.duplicate",
                    f"row:{row.source_row_number}:order_id+order_line_id",
                    "canonical_dictionary:30",
                    "duplicate order_id + order_line_id creates unresolved row identity ambiguity",
                )
            )
        identities[identity] = row.source_row_number
        if row.order_id in order_dates and order_dates[row.order_id] != row.order_date:
            inconsistent_order_dates = True
            results.append(
                blocking(
                    "canonical.order_date.inconsistent_within_order",
                    row.order_id,
                    "canonical_dictionary:16",
                    "all lines with the same order_id must have one consistent order_date",
                )
            )
        order_dates[row.order_id] = row.order_date
        if row.product_name:
            product_names.setdefault(row.product_id, set()).add(row.product_name)
        currencies.add(row.currency)
    if not duplicate_identity:
        results.append(
            passed(
                "canonical.identity.unique",
                "order_id+order_line_id",
                "canonical_dictionary:7.1",
                "composite order-line identity is unique",
            )
        )
    if not inconsistent_order_dates:
        results.append(
            passed(
                "canonical.order_date.consistent_within_order",
                "order_id",
                "canonical_dictionary:16",
                "each order has one consistent order_date",
            )
        )
    if len(currencies) > 1:
        results.append(
            blocking(
                "canonical.currency.mixed",
                "currency",
                "canonical_dictionary:32.2",
                "multiple unnormalized currencies block monetary comparison",
            )
        )
    elif request.currency_basis_ref is not None and currencies != {request.currency_basis_ref}:
        results.append(
            blocking(
                "canonical.currency.basis_mismatch",
                "currency",
                "canonical_dictionary:32.2",
                "canonical row currency does not match the explicit governed currency basis",
            )
        )
    else:
        results.append(
            passed(
                "canonical.currency.single_basis",
                "currency",
                "canonical_dictionary:32.2",
                "canonical rows use one explicit currency basis",
            )
        )
    for product_id, names in sorted(product_names.items()):
        if len(names) > 1:
            results.append(
                qualification(
                    "canonical.product_name.changed",
                    product_id,
                    "canonical_dictionary:19",
                    "product_name changed while product_id remains the authoritative identity",
                )
            )
    if request.require_category and any(row.category_id == UNCLASSIFIED_CATEGORY_ID for row in rows):
        results.append(
            qualification(
                "canonical.category.contains_unclassified",
                UNCLASSIFIED_CATEGORY_ID,
                "canonical_dictionary:20",
                "category analysis contains the governed Unclassified bucket",
            )
        )
    return tuple(results)


def _qualifications_from_quality(results: list[DataQualityCheckResult]) -> tuple[Qualification, ...]:
    qualifications: list[Qualification] = []
    seen: set[tuple[str, str]] = set()
    for result in results:
        if result.consequence is not DataQualityConsequence.QUALIFICATION:
            continue
        key = (result.check_id, result.target)
        if key in seen:
            continue
        seen.add(key)
        qualifications.append(
            Qualification(
                qualification_id=stable_content_id(
                    "qual",
                    canonical_json_fingerprint({"check_id": result.check_id, "target": result.target}),
                ),
                statement=result.reason,
                affected_ref=result.target,
            )
        )
    return tuple(qualifications)


def _blocked_result(
    dataset: DatasetReference,
    request: CanonicalizationRequest,
    quality_results: list[DataQualityCheckResult],
    *,
    qualifications: tuple[Qualification, ...] = (),
) -> CanonicalizationResult:
    failures = tuple(
        result.reason for result in quality_results if result.consequence is DataQualityConsequence.BLOCKING
    )
    record = _record(
        dataset=dataset,
        request=request,
        canonical_dataset_id=None,
        output_fingerprint=None,
        source_row_count=None,
        canonical_row_count=None,
        quality_results=quality_results,
        qualifications=qualifications,
        failures=failures,
        selected_sheet=None,
        selected_table=None,
    )
    return CanonicalizationResult(
        record=record,
        data_quality_results=tuple(quality_results),
        qualifications=qualifications,
        failures=failures,
    )


def _record(
    *,
    dataset: DatasetReference,
    request: CanonicalizationRequest,
    canonical_dataset_id: str | None,
    output_fingerprint: str | None,
    source_row_count: int | None,
    canonical_row_count: int | None,
    quality_results: list[DataQualityCheckResult],
    qualifications: tuple[Qualification, ...],
    failures: tuple[str, ...],
    selected_sheet: str | None,
    selected_table: str | None,
) -> CanonicalizationRecord:
    mapping_fingerprint = request.mapping.fingerprint
    canonicalization_fingerprint = canonical_json_fingerprint(
        {
            "material_config": _material_config(
                dataset,
                request,
                mapping_fingerprint,
                selected_sheet=selected_sheet,
                selected_table=selected_table,
            ),
            "source_row_count": source_row_count,
            "canonical_row_count": canonical_row_count,
            "quality_results": [result.model_dump(mode="json") for result in quality_results],
            "failures": failures,
        }
    )
    return CanonicalizationRecord(
        canonicalization_id=stable_content_id("canonrec", canonicalization_fingerprint),
        source_dataset_id=dataset.dataset_id,
        canonical_dataset_id=canonical_dataset_id,
        canonical_schema_version=request.canonical_schema_version,
        mapping_ref=request.mapping.mapping_id,
        transformation_version=CANONICALIZATION_VERSION,
        mapping_fingerprint=mapping_fingerprint,
        eligibility_mode=request.eligibility_mode.value,
        selected_sheet=selected_sheet,
        selected_table=selected_table,
        date_policy_ref=request.date_policy_ref,
        currency_basis_ref=request.currency_basis_ref,
        unsupported_partial_refund_evidence=request.unsupported_partial_refund_evidence,
        source_fingerprint=dataset.content_fingerprint,
        output_fingerprint=output_fingerprint,
        source_row_count=source_row_count,
        canonical_row_count=canonical_row_count,
        data_quality_result_ids=tuple(result.check_id for result in quality_results),
        qualifications=tuple(qualification.statement for qualification in qualifications),
        failures=failures,
    )


def _write_canonical_artifact(
    dataset: DatasetReference,
    request: CanonicalizationRequest,
    rows: list[CanonicalLineRecord],
    artifact_store: ArtifactStore,
    *,
    selected_sheet: str | None,
    selected_table: str | None,
) -> CanonicalDatasetReference:
    json_rows = [row.model_dump(mode="json") for row in rows]
    semantic_fingerprint = canonical_json_fingerprint(
        {
            "material_config": _material_config(
                dataset,
                request,
                request.mapping.fingerprint,
                selected_sheet=selected_sheet,
                selected_table=selected_table,
            ),
            "rows": json_rows,
        }
    )
    canonical_dataset_id = stable_content_id("cds", semantic_fingerprint)
    artifact_store.ensure_layout()
    artifact_path = artifact_store.safe_path("canonical", semantic_fingerprint[:2], f"{canonical_dataset_id}.parquet")
    artifact_path.parent.mkdir(parents=True, exist_ok=True)
    _write_parquet(rows, artifact_path)
    artifact_fingerprint = sha256_file(artifact_path)
    artifact = ArtifactReference(
        artifact_id=stable_content_id("art", artifact_fingerprint),
        path=str(artifact_path.relative_to(artifact_store.root)),
        fingerprint=artifact_fingerprint,
        media_type="application/vnd.apache.parquet",
        size_bytes=artifact_path.stat().st_size,
    )
    return CanonicalDatasetReference(
        canonical_dataset_id=canonical_dataset_id,
        source_dataset_id=dataset.dataset_id,
        canonical_schema_version=request.canonical_schema_version,
        content_fingerprint=artifact_fingerprint,
        artifact=artifact,
        row_count=len(rows),
    )


def _material_config(
    dataset: DatasetReference,
    request: CanonicalizationRequest,
    mapping_fingerprint: str,
    *,
    selected_sheet: str | None,
    selected_table: str | None,
) -> dict[str, Any]:
    return {
        "source_dataset_id": dataset.dataset_id,
        "source_fingerprint": dataset.content_fingerprint,
        "selected_sheet": selected_sheet,
        "selected_table": selected_table,
        "canonical_schema_version": request.canonical_schema_version,
        "mapping_ref": request.mapping.mapping_id,
        "mapping_fingerprint": mapping_fingerprint,
        "eligibility_mode": request.eligibility_mode.value,
        "eligibility_mapping": [entry.model_dump(mode="json") for entry in request.eligibility_value_mapping],
        "require_category": request.require_category,
        "date_policy_ref": request.date_policy_ref,
        "currency_basis_ref": request.currency_basis_ref,
        "unsupported_partial_refund_evidence": request.unsupported_partial_refund_evidence,
        "transformation_version": CANONICALIZATION_VERSION,
    }


def _monetary_artifact_quality(rows: list[CanonicalLineRecord]) -> tuple[DataQualityCheckResult, ...]:
    results: list[DataQualityCheckResult] = []
    for field_name, values in (
        ("line_revenue", [row.line_revenue for row in rows]),
        ("unit_price", [row.unit_price for row in rows if row.unit_price is not None]),
    ):
        try:
            _duckdb_decimal_type(values)
        except ValueError:
            results.append(
                blocking(
                    f"canonical.artifact.{field_name}.unrepresentable_precision",
                    field_name,
                    "canonical_dictionary:32.1",
                    f"{field_name} cannot be represented exactly within DuckDB DECIMAL precision",
                )
            )
    return tuple(results)


def _duckdb_decimal_type(values: list[Decimal]) -> str:
    if not values:
        return "DECIMAL(1, 0)"
    max_scale = 0
    max_integer_digits = 1
    for value in values:
        precision, scale = _decimal_precision_scale(value)
        if precision > 38:
            raise ValueError("Decimal value cannot be represented exactly within DuckDB DECIMAL precision")
        max_scale = max(max_scale, scale)
        max_integer_digits = max(max_integer_digits, precision - scale)
    width = max_integer_digits + max_scale
    if width > 38:
        raise ValueError("Decimal values require DuckDB DECIMAL width greater than 38")
    return f"DECIMAL({width}, {max_scale})"


def _decimal_precision_scale(value: Decimal) -> tuple[int, int]:
    if not value.is_finite():
        return 39, 0
    sign, digits, exponent = value.as_tuple()
    del sign
    digit_count = len(digits)
    if exponent >= 0:
        return digit_count + exponent, 0
    scale = -exponent
    integer_digits = max(0, digit_count - scale)
    return max(integer_digits + scale, 1), scale


def _write_parquet(rows: list[CanonicalLineRecord], artifact_path: Path) -> None:
    line_revenue_type = _duckdb_decimal_type([row.line_revenue for row in rows])
    unit_price_type = _duckdb_decimal_type([row.unit_price for row in rows if row.unit_price is not None])
    conn = duckdb.connect(database=":memory:")
    try:
        conn.execute(
            f"""
            CREATE TABLE canonical_lines (
                source_row_number BIGINT,
                order_id VARCHAR,
                order_line_id VARCHAR,
                order_date DATE,
                product_id VARCHAR,
                product_name VARCHAR,
                category_id VARCHAR,
                category_name VARCHAR,
                quantity BIGINT,
                line_revenue {line_revenue_type},
                currency VARCHAR,
                eligibility_status VARCHAR,
                unit_price {unit_price_type}
            )
            """
        )
        conn.executemany(
            """
            INSERT INTO canonical_lines VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row.source_row_number,
                    row.order_id,
                    row.order_line_id,
                    row.order_date,
                    row.product_id,
                    row.product_name,
                    row.category_id,
                    row.category_name,
                    row.quantity,
                    row.line_revenue,
                    row.currency,
                    row.eligibility_status.value,
                    row.unit_price,
                )
                for row in rows
            ],
        )
        conn.execute("COPY canonical_lines TO ? (FORMAT PARQUET)", (str(artifact_path),))
    finally:
        conn.close()
