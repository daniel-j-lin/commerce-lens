from __future__ import annotations

import json
import subprocess
import sys
from dataclasses import replace
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from commerce_lens.contracts.common import ClaimState, MetricState, PeriodDefinition, SourceType
from commerce_lens.persistence.artifact_store import ArtifactStore
from commerce_lens.persistence.metadata_store import MetadataStore
from commerce_lens.skill.integration import (
    PublicAnalysisIntent,
    PublicQuestionClass,
    PublicSourceSelection,
    run_public_analysis,
)
from commerce_lens.skill.schema_mapping import assess_schema_mapping, confirmed_mapping_from_source_to_canonical


ROOT = Path(__file__).resolve().parents[2]
FIXTURES = ROOT / "tests" / "fixtures" / "p14"
RUNNER = ROOT / "skills" / "commerce-lens" / "scripts" / "run_public_analysis.py"

EXPECTED_REVENUE_CHANGE = Decimal("-20.00")
EXPECTED_Q4_REVENUE = Decimal("100.00")
EXPECTED_Q4_ORDERS = 2
EXPECTED_Q4_AOV = Decimal("50.00")

CANONICAL_REQUIRED = (
    "order_id",
    "order_line_id",
    "order_date",
    "product_id",
    "quantity",
    "line_revenue",
    "currency",
    "eligibility_status",
)

SOURCE_TO_CANONICAL: dict[str, dict[str, str] | None] = {
    "P14-A": None,
    "P14-B": {
        "Order Number": "order_id",
        "Line Item ID": "order_line_id",
        "Order Date": "order_date",
        "SKU": "product_id",
        "Product": "product_name",
        "Category": "category_name",
        "Quantity": "quantity",
        "Sales Amount": "line_revenue",
        "Currency": "currency",
        "Unit Price": "unit_price",
        "Order Status": "eligibility_status",
    },
    "P14-C": {
        "Name": "order_id",
        "Lineitem id": "order_line_id",
        "Created at": "order_date",
        "Lineitem sku": "product_id",
        "Lineitem name": "product_name",
        "Product type": "category_name",
        "Lineitem quantity": "quantity",
        "Lineitem total": "line_revenue",
        "Currency": "currency",
        "Lineitem price": "unit_price",
        "Financial Status": "eligibility_status",
    },
    "P14-D": {
        "Order ID": "order_id",
        "Order Item ID": "order_line_id",
        "Date Created": "order_date",
        "Product ID": "product_id",
        "Product Name": "product_name",
        "Qty": "quantity",
        "Line Total": "line_revenue",
        "Currency": "currency",
        "Status": "eligibility_status",
    },
    "P14-E": {
        "Document No": "order_id",
        "Line No": "order_line_id",
        "Posting Date": "order_date",
        "Item Code": "product_id",
        "Item Description": "product_name",
        "Qty": "quantity",
        "Net Amount": "line_revenue",
        "Currency Code": "currency",
        "Unit Amount": "unit_price",
        "Document Status": "eligibility_status",
    },
    "P14-F": {
        " ORDER number ": "order_id",
        "line item id": "order_line_id",
        "Posted Date": "order_date",
        "Item SKU": "product_id",
        "Item Name": "product_name",
        " Qty ": "quantity",
        " Sales Amount ": "line_revenue",
        "Currency Code": "currency",
        "Each Price": "unit_price",
        " Order Status ": "eligibility_status",
    },
    "P14-G": {
        "Order Number": "order_id",
        "Line Item ID": "order_line_id",
        "Order Date": "order_date",
        "SKU": "product_id",
        "Product": "product_name",
        "Category": "category_name",
        "Quantity": "quantity",
        "Sales Amount": "line_revenue",
        "Currency": "currency",
        "Unit Price": "unit_price",
        "Order Status": "eligibility_status",
    },
    "P14-H": {
        "Order ID": "order_id",
        "Order Line ID": "order_line_id",
        "Order Date": "order_date",
        "SKU": "product_id",
        "Quantity": "quantity",
        "Total": "line_revenue",
        "Currency": "currency",
        "Status": "eligibility_status",
    },
}

FILES = {
    "P14-A": ("P14-A-canonical-control.csv", SourceType.CSV),
    "P14-B": ("P14-B-generic-marketplace.csv", SourceType.CSV),
    "P14-C": ("P14-C-shopify-like-synthetic.csv", SourceType.CSV),
    "P14-D": ("P14-D-woocommerce-like-synthetic.csv", SourceType.CSV),
    "P14-E": ("P14-E-erp-back-office.csv", SourceType.CSV),
    "P14-F": ("P14-F-messy-but-valid.csv", SourceType.CSV),
    "P14-G": ("P14-G-xlsx-native-export.xlsx", SourceType.EXCEL_XLSX),
    "P14-H": ("P14-H-insufficient-unsupported.csv", SourceType.CSV),
    "P14-H2": ("P14-H2-order-grain-insufficient.csv", SourceType.CSV),
}


def test_p14_fixture_inventory_files_exist() -> None:
    assert set(FILES) == {"P14-A", "P14-B", "P14-C", "P14-D", "P14-E", "P14-F", "P14-G", "P14-H", "P14-H2"}
    for filename, _source_type in FILES.values():
        assert (FIXTURES / filename).is_file()


def test_p14_xlsx_fixture_uses_native_typed_cells() -> None:
    workbook = load_workbook(FIXTURES / FILES["P14-G"][0], data_only=True)
    sheet = workbook["Orders"]

    assert isinstance(sheet["C2"].value, datetime)
    assert isinstance(sheet["H2"].value, int)
    assert sheet["C2"].number_format == "yyyy-mm-dd"
    assert sheet["H2"].number_format == "$#,##0.00"


def test_p14_canonical_control_needs_no_mapping_confirmation(tmp_path: Path) -> None:
    outcome = _run(tmp_path, _intent("P14-A", "revenue_change"))

    assert outcome.response.clarification_required == ()
    assert outcome.response.mapping_proposals == ()
    claim = outcome.response.supported_claims[0]
    assert claim.metric_ref == "revenue_change"
    assert claim.value == EXPECTED_REVENUE_CHANGE
    assert claim.claim_state is ClaimState.ADMISSIBLE


def test_p14_noncanonical_fixtures_block_before_explicit_mapping_confirmation(tmp_path: Path) -> None:
    for fixture_id in ("P14-B", "P14-C", "P14-D", "P14-E", "P14-F", "P14-G", "P14-H"):
        outcome = _run(tmp_path / fixture_id, _intent(fixture_id, "revenue_change"))

        assert outcome.request is None
        assert outcome.response.blocked is True
        assert outcome.response.supported_claims == ()
        assert outcome.response.clarification_required


def test_p14_successful_confirmed_fixtures_are_canonically_equivalent(tmp_path: Path) -> None:
    for fixture_id in ("P14-A", "P14-B", "P14-C", "P14-D", "P14-E", "P14-F", "P14-G"):
        change = _run(tmp_path / fixture_id / "change", _intent(fixture_id, "revenue_change", confirmed=True))
        revenue = _run(tmp_path / fixture_id / "revenue", _intent(fixture_id, "revenue", confirmed=True))
        orders = _run(tmp_path / fixture_id / "orders", _intent(fixture_id, "orders", confirmed=True))
        aov = _run(tmp_path / fixture_id / "aov", _intent(fixture_id, "aov", confirmed=True))

        assert _single_claim(change).value == EXPECTED_REVENUE_CHANGE
        assert _single_claim(revenue).value == EXPECTED_Q4_REVENUE
        assert _single_claim(orders).value == EXPECTED_Q4_ORDERS
        assert _single_claim(aov).value == EXPECTED_Q4_AOV
        assert _single_claim(change).claim_state is ClaimState.ADMISSIBLE
        assert change.response.evidence_summary[0].validation_status == "passed"


def test_p14_eligibility_authority_excludes_cancelled_revenue(tmp_path: Path) -> None:
    revenue = _run(tmp_path / "revenue", _intent("P14-B", "revenue", confirmed=True))
    orders = _run(tmp_path / "orders", _intent("P14-B", "orders", confirmed=True))

    assert _single_claim(revenue).value == EXPECTED_Q4_REVENUE
    assert _single_claim(orders).value == EXPECTED_Q4_ORDERS
    assert _single_claim(revenue).value != Decimal("1099.00")
    assert _single_claim(orders).value != 3


def test_p14_insufficient_fixture_fails_closed_on_unresolved_eligibility_semantics(tmp_path: Path) -> None:
    outcome = _run(tmp_path, _intent("P14-H", "revenue_change", confirmed=True))

    assert outcome.response.blocked is True
    assert outcome.response.supported_claims == ()
    assert outcome.response.insufficient_evidence_message == "Insufficient evidence to conclude."
    assert any("source eligibility value is not explicitly mapped" in item for item in outcome.response.limitations)


def test_p14_order_grain_export_fails_closed_without_line_level_authority() -> None:
    headers = _headers_for("P14-H2")
    assessment = assess_schema_mapping(headers, "revenue_change")

    assert assessment.identity_mapping_available is False
    assert "order_line_id" in assessment.missing_required_fields
    assert "quantity" in assessment.missing_required_fields
    assert "Confirm or correct" in " ".join(assessment.clarification_required)


def test_p14_skill_runner_e2e_representative_fixtures() -> None:
    required = ("P14-B", "P14-C", "P14-E", "P14-G", "P14-H")
    for fixture_id in required:
        payload = _run_runner(fixture_id, confirmed=True)

        if fixture_id == "P14-H":
            assert payload["response"]["blocked"] is True
            assert payload["response"]["supported_claims"] == []
            assert "source eligibility value is not explicitly mapped" in " ".join(payload["response"]["limitations"])
        else:
            claim = payload["response"]["supported_claims"][0]
            assert claim["metric_ref"] == "revenue_change"
            assert Decimal(claim["value"]) == EXPECTED_REVENUE_CHANGE
            assert claim["claim_state"] == "Admissible"
            assert payload["validated_results_summary"]


def test_p14_skill_runner_blocks_unconfirmed_material_mapping() -> None:
    payload = _run_runner("P14-B", confirmed=False)

    assert payload["request_id"] is None
    assert payload["response"]["blocked"] is True
    assert payload["response"]["supported_claims"] == []
    assert payload["response"]["mapping_proposals"]


def _intent(fixture_id: str, metric: str, *, confirmed: bool = False) -> PublicAnalysisIntent:
    filename, source_type = FILES[fixture_id]
    source_path = FIXTURES / filename
    result_period_role = "comparison" if metric in {"revenue", "orders", "aov"} else None
    question_class = (
        PublicQuestionClass.SINGLE_PERIOD_METRIC
        if metric in {"revenue", "orders", "aov"}
        else PublicQuestionClass.REVENUE_CHANGE
    )
    source = PublicSourceSelection(
        source_path,
        source_type,
        selected_sheet="Orders" if source_type is SourceType.EXCEL_XLSX else None,
    )
    if confirmed and SOURCE_TO_CANONICAL.get(fixture_id) is not None:
        source = replace(
            source,
            mapping=confirmed_mapping_from_source_to_canonical(SOURCE_TO_CANONICAL[fixture_id] or {}),
            mapping_mode="confirmed_source_to_canonical_mapping",
        )
    return PublicAnalysisIntent(
        question_class=question_class,
        metric_id=metric,
        baseline_period=PeriodDefinition(
            period_id="baseline",
            label="Q3 2026",
            start_date=date(2026, 7, 1),
            end_date=date(2026, 9, 30),
            date_convention_ref="order_date_utc",
        ),
        comparison_period=PeriodDefinition(
            period_id="comparison",
            label="Q4 2026",
            start_date=date(2026, 10, 1),
            end_date=date(2026, 12, 31),
            date_convention_ref="order_date_utc",
        ),
        source=source,
        original_question_text=(
            "How did revenue change from Q3 2026 to Q4 2026?"
            if metric == "revenue_change"
            else f"What was {metric} in Q4 2026?"
        ),
        result_period_role=result_period_role,
    )


def _run(tmp_path: Path, intent: PublicAnalysisIntent):
    return run_public_analysis(
        intent,
        artifact_store=ArtifactStore(tmp_path / "artifacts"),
        metadata_store=MetadataStore(tmp_path / "metadata.sqlite"),
    )


def _single_claim(outcome):
    assert len(outcome.response.supported_claims) == 1
    claim = outcome.response.supported_claims[0]
    assert claim.metric_state is MetricState.VALID
    return claim


def _headers_for(fixture_id: str) -> tuple[str, ...]:
    filename, source_type = FILES[fixture_id]
    if source_type is SourceType.EXCEL_XLSX:
        workbook = load_workbook(FIXTURES / filename, data_only=True, read_only=True)
        return tuple(str(value).strip() if value is not None else "" for value in next(workbook["Orders"].iter_rows(values_only=True, max_row=1)))
    with (FIXTURES / filename).open("r", encoding="utf-8-sig", newline="") as file_obj:
        return tuple(file_obj.readline().rstrip("\n").split(","))


def _run_runner(fixture_id: str, *, confirmed: bool) -> dict:
    filename, source_type = FILES[fixture_id]
    args = [
        sys.executable,
        str(RUNNER),
        "--source",
        str(FIXTURES / filename),
        "--source-type",
        "xlsx" if source_type is SourceType.EXCEL_XLSX else "csv",
        "--question-class",
        "revenue_change",
        "--metric",
        "revenue_change",
        "--baseline-label",
        "Q3 2026",
        "--baseline-start",
        "2026-07-01",
        "--baseline-end",
        "2026-09-30",
        "--comparison-label",
        "Q4 2026",
        "--comparison-start",
        "2026-10-01",
        "--comparison-end",
        "2026-12-31",
        "--original-question",
        "How did revenue change from Q3 2026 to Q4 2026?",
    ]
    if source_type is SourceType.EXCEL_XLSX:
        args.extend(["--selected-sheet", "Orders"])
    if confirmed and SOURCE_TO_CANONICAL.get(fixture_id) is not None:
        args.extend(["--mapping-json", json.dumps(SOURCE_TO_CANONICAL[fixture_id])])

    completed = subprocess.run(args, cwd=ROOT, text=True, capture_output=True, check=False)
    assert completed.returncode == 0, completed.stderr
    return json.loads(completed.stdout)
