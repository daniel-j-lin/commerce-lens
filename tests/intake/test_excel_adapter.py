from openpyxl import Workbook

from commerce_lens.evidence.identifiers import sha256_file
from commerce_lens.intake.excel_adapter import ExcelInspectionAdapter
from commerce_lens.intake.inspection import InspectionStatus


def test_xlsx_inspection_requires_explicit_sheet_when_ambiguous(tmp_path) -> None:
    path = tmp_path / "orders.xlsx"
    workbook = Workbook()
    workbook.active.title = "January"
    workbook.create_sheet("February")
    workbook.save(path)

    result = ExcelInspectionAdapter().inspect(path)

    assert result.status is InspectionStatus.AMBIGUOUS
    assert result.available_sheets == ("January", "February")


def test_xlsx_inspection_discovers_selected_sheet_and_preserves_source(tmp_path) -> None:
    path = tmp_path / "orders.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(["order_id", "line_revenue"])
    sheet.append(["1", 10.5])
    workbook.save(path)
    before = sha256_file(path)

    result = ExcelInspectionAdapter().inspect(path, sheet_name="Orders")

    assert result.status is InspectionStatus.SUPPORTED
    assert result.selected_sheet == "Orders"
    assert [column.name for column in result.columns] == ["order_id", "line_revenue"]
    assert result.row_count == 1
    assert sha256_file(path) == before


def test_xls_is_unsupported(tmp_path) -> None:
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"not really excel")

    result = ExcelInspectionAdapter().inspect(path)

    assert result.status is InspectionStatus.UNSUPPORTED


def test_formula_cell_without_stored_value_fails_closed_and_preserves_source(tmp_path) -> None:
    path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(["order_id", "line_revenue"])
    sheet.append(["1", "=1+2"])
    workbook.save(path)
    before = sha256_file(path)

    result = ExcelInspectionAdapter().inspect(path, sheet_name="Orders")

    assert result.status is InspectionStatus.FAILED
    assert "has no stored value" in result.failure_detail.reason
    assert "does not recalculate" in result.failure_detail.reason
    assert sha256_file(path) == before
