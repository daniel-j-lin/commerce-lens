import csv
import sqlite3
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import duckdb
from openpyxl import Workbook

from commerce_lens.canonical import (
    UNCLASSIFIED_CATEGORY_ID,
    CanonicalFieldMapping,
    CanonicalMapping,
    CanonicalizationRequest,
    EligibilityMode,
    EligibilityState,
    EligibilityValueMapping,
    canonicalize_dataset,
    identity_mapping,
)
from commerce_lens.canonical.quality import DataQualityConsequence
from commerce_lens.contracts.common import SourceType
from commerce_lens.evidence.identifiers import sha256_file
from commerce_lens.intake.registry import DatasetRegistry
from commerce_lens.persistence.artifact_store import ArtifactStore


def test_valid_multiline_order_and_same_product_separate_lines_are_canonicalized(tmp_path) -> None:
    dataset, store, before = _registered_csv(
        tmp_path,
        [
            {
                "order_id": "o1",
                "order_line_id": "l1",
                "order_date": "2026-01-01",
                "product_id": "p1",
                "product_name": "Tea",
                "quantity": "1",
                "line_revenue": "0",
                "currency": "USD",
            },
            {
                "order_id": "o1",
                "order_line_id": "l2",
                "order_date": "2026-01-01",
                "product_id": "p1",
                "product_name": "Tea",
                "quantity": "2",
                "line_revenue": "9.50",
                "currency": "USD",
            },
        ],
    )

    result = canonicalize_dataset(dataset, _request(dataset, require_category=False), store)

    assert result.canonical_dataset is not None
    assert [row.order_line_id for row in result.canonical_rows] == ["l1", "l2"]
    assert result.canonical_rows[0].line_revenue == 0
    assert sha256_file(tmp_path / "orders.csv") == before
    parquet_path = store.safe_path(result.canonical_dataset.artifact.path)
    assert parquet_path.exists()
    assert duckdb.sql(f"SELECT count(*) FROM read_parquet('{parquet_path}')").fetchone()[0] == 2


def test_duplicate_order_line_identity_fails_without_deduplication(tmp_path) -> None:
    dataset, store, _ = _registered_csv(
        tmp_path,
        [
            _row(order_id="o1", order_line_id="l1"),
            _row(order_id="o1", order_line_id="l1", line_revenue="12.00"),
        ],
    )

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert result.canonical_dataset is None
    assert any(check.check_id == "canonical.identity.duplicate" for check in result.data_quality_results)


def test_missing_identity_fields_fail(tmp_path) -> None:
    dataset, store, _ = _registered_csv(tmp_path, [_row(order_id="")])

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert any(check.check_id == "canonical.order_id.missing" for check in result.data_quality_results)
    assert result.failures


def test_ambiguous_date_string_and_inconsistent_order_dates_fail(tmp_path) -> None:
    dataset, store, _ = _registered_csv(
        tmp_path,
        [
            _row(order_id="o1", order_line_id="l1", order_date="01/02/2026"),
            _row(order_id="o2", order_line_id="l1", order_date="2026-01-01"),
            _row(order_id="o2", order_line_id="l2", order_date="2026-01-02"),
        ],
    )

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert any(check.check_id == "canonical.order_date.invalid" for check in result.data_quality_results)
    assert any(
        check.check_id == "canonical.order_date.inconsistent_within_order"
        for check in result.data_quality_results
    )


def test_product_identity_uses_id_not_name_and_changed_name_qualifies(tmp_path) -> None:
    dataset, store, _ = _registered_csv(
        tmp_path,
        [
            _row(order_id="o1", product_id="p1", product_name="Old"),
            _row(order_id="o2", product_id="p1", product_name="New"),
            _row(order_id="o3", product_id="p2", product_name="New"),
        ],
    )

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert result.canonical_dataset is not None
    assert {row.product_id for row in result.canonical_rows} == {"p1", "p2"}
    assert any(check.check_id == "canonical.product_name.changed" for check in result.data_quality_results)


def test_missing_product_id_blocks_without_product_name_substitution(tmp_path) -> None:
    dataset, store, _ = _registered_csv(tmp_path, [_row(product_id="", product_name="Known Name")])

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert any(check.check_id == "canonical.product_id.missing" for check in result.data_quality_results)
    assert result.canonical_dataset is None


def test_missing_category_becomes_visible_unclassified_when_category_required(tmp_path) -> None:
    dataset, store, _ = _registered_csv(tmp_path, [_row(category_id="", category_name="")])

    result = canonicalize_dataset(dataset, _request(dataset, require_category=True), store)

    assert result.canonical_dataset is not None
    assert result.canonical_rows[0].category_id == UNCLASSIFIED_CATEGORY_ID
    assert result.canonical_rows[0].category_name == "Unclassified"
    assert any(check.check_id == "canonical.category.contains_unclassified" for check in result.data_quality_results)


def test_quantity_invalid_cases_fail_closed(tmp_path) -> None:
    for quantity in ("0", "-1", "1.5", "", "true"):
        dataset, store, _ = _registered_csv(tmp_path, [_row(quantity=quantity)], filename=f"q_{quantity or 'blank'}.csv")

        result = canonicalize_dataset(dataset, _request(dataset), store)

        assert any(check.check_id == "canonical.quantity.invalid" for check in result.data_quality_results)
        assert result.canonical_dataset is None


def test_revenue_missing_negative_and_float_style_values_fail_but_zero_is_valid(tmp_path) -> None:
    dataset, store, _ = _registered_csv(tmp_path, [_row(line_revenue="0")], filename="zero.csv")
    assert canonicalize_dataset(dataset, _request(dataset), store).canonical_dataset is not None

    for value in ("", "-1.00"):
        dataset, store, _ = _registered_csv(tmp_path, [_row(line_revenue=value)], filename=f"revenue_{value or 'blank'}.csv")
        result = canonicalize_dataset(dataset, _request(dataset), store)
        assert any("line_revenue" in check.check_id for check in result.data_quality_results)
        assert result.canonical_dataset is None


def test_quantity_times_unit_price_does_not_fill_missing_line_revenue(tmp_path) -> None:
    dataset, store, _ = _registered_csv(tmp_path, [_row(line_revenue="", unit_price="5.00", quantity="2")])

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert result.canonical_dataset is None
    assert any(check.check_id == "canonical.line_revenue.invalid" for check in result.data_quality_results)


def test_currency_single_basis_passes_and_mixed_or_missing_blocks(tmp_path) -> None:
    dataset, store, _ = _registered_csv(tmp_path, [_row(currency="USD")], filename="usd.csv")
    assert canonicalize_dataset(dataset, _request(dataset), store).canonical_dataset is not None

    dataset, store, _ = _registered_csv(tmp_path, [_row(order_id="o1", currency="USD"), _row(order_id="o2", currency="EUR")], filename="mixed.csv")
    mixed = canonicalize_dataset(dataset, _request(dataset), store)
    assert any(check.check_id == "canonical.currency.mixed" for check in mixed.data_quality_results)

    dataset, store, _ = _registered_csv(tmp_path, [_row(currency="")], filename="missing_currency.csv")
    missing = canonicalize_dataset(dataset, _request(dataset), store)
    assert any(check.check_id == "canonical.currency.missing" for check in missing.data_quality_results)


def test_explicit_eligibility_mapping_retains_excluded_and_unknown_status_fails(tmp_path) -> None:
    dataset, store, _ = _registered_csv(
        tmp_path,
        [_row(order_id="o1", eligibility_status="paid"), _row(order_id="o2", eligibility_status="cancelled")],
        include_status=True,
    )

    result = canonicalize_dataset(dataset, _request(dataset, eligibility=True), store)

    assert [row.eligibility_status for row in result.canonical_rows] == [
        EligibilityState.ELIGIBLE,
        EligibilityState.EXCLUDED,
    ]

    dataset, store, _ = _registered_csv(
        tmp_path,
        [_row(eligibility_status="mystery")],
        filename="unknown_status.csv",
        include_status=True,
    )
    unknown = canonicalize_dataset(dataset, _request(dataset, eligibility=True), store)
    assert any(check.check_id == "canonical.eligibility.unknown" for check in unknown.data_quality_results)


def test_partial_refund_evidence_blocks_without_reconstruction(tmp_path) -> None:
    dataset, store, _ = _registered_csv(tmp_path, [_row()])

    result = canonicalize_dataset(dataset, _request(dataset, partial_refund=True), store)

    assert result.canonical_dataset is None
    assert any(check.check_id == "canonical.partial_refund_unsupported" for check in result.data_quality_results)


def _request(
    dataset,
    *,
    require_category: bool = False,
    eligibility: bool = False,
    partial_refund: bool = False,
) -> CanonicalizationRequest:
    headers = _headers(include_status=eligibility)
    mapping = identity_mapping(headers, require_category=require_category, require_eligibility=eligibility)
    return CanonicalizationRequest(
        source_dataset_id=dataset.dataset_id,
        mapping=mapping,
        eligibility_mode=EligibilityMode.EXPLICIT_STATUS_MAPPING if eligibility else EligibilityMode.UPSTREAM_ELIGIBLE_ONLY,
        eligibility_value_mapping=(
            (
                EligibilityValueMapping(source_value="paid", normalized_status=EligibilityState.ELIGIBLE),
                EligibilityValueMapping(source_value="cancelled", normalized_status=EligibilityState.EXCLUDED),
            )
            if eligibility
            else ()
        ),
        require_category=require_category,
        unsupported_partial_refund_evidence=partial_refund,
    )


def _registered_csv(tmp_path: Path, rows: list[dict[str, str]], *, filename: str = "orders.csv", include_status: bool = False):
    path = tmp_path / filename
    headers = _headers(include_status=include_status)
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=headers)
        writer.writeheader()
        writer.writerows([{key: row.get(key, "") for key in headers} for row in rows])
    before = sha256_file(path)
    store = ArtifactStore(tmp_path / f"runtime_{filename}")
    registry = DatasetRegistry(store)
    return registry.register_source(path, SourceType.CSV), store, before


def _headers(*, include_status: bool = False) -> tuple[str, ...]:
    headers = (
        "order_id",
        "order_line_id",
        "order_date",
        "product_id",
        "product_name",
        "category_id",
        "category_name",
        "quantity",
        "line_revenue",
        "currency",
        "unit_price",
    )
    return headers + (("eligibility_status",) if include_status else ())


def _row(**overrides: str) -> dict[str, str]:
    row = {
        "order_id": "o1",
        "order_line_id": "l1",
        "order_date": "2026-01-01",
        "product_id": "p1",
        "product_name": "Tea",
        "category_id": "c1",
        "category_name": "Drinks",
        "quantity": "1",
        "line_revenue": "10.00",
        "currency": "USD",
        "unit_price": "",
        "eligibility_status": "paid",
    }
    row.update(overrides)
    return row


def test_csv_xlsx_and_sqlite_monetary_values_converge(tmp_path) -> None:
    rows = [_row(line_revenue="10.25", unit_price="5.125", quantity="2")]
    csv_dataset, csv_store, _ = _registered_csv(tmp_path, rows, filename="converge.csv")
    xlsx_dataset, xlsx_store = _registered_xlsx(tmp_path, rows, numeric=True)
    sqlite_dataset, sqlite_store = _registered_sqlite(tmp_path, rows, numeric=True)

    results = [
        canonicalize_dataset(csv_dataset, _request(csv_dataset), csv_store),
        canonicalize_dataset(xlsx_dataset, _request(xlsx_dataset), xlsx_store),
        canonicalize_dataset(sqlite_dataset, _request(sqlite_dataset), sqlite_store),
    ]

    material_values = [
        (
            result.canonical_rows[0].quantity,
            result.canonical_rows[0].line_revenue,
            result.canonical_rows[0].unit_price,
            result.canonical_rows[0].currency,
        )
        for result in results
    ]
    assert material_values == [(2, Decimal("10.25"), Decimal("5.125"), "USD")] * 3


def test_non_finite_monetary_and_quantity_inputs_fail_closed(tmp_path) -> None:
    for field_name, value in (("line_revenue", "NaN"), ("line_revenue", "Infinity"), ("quantity", "NaN")):
        dataset, store, _ = _registered_csv(
            tmp_path,
            [_row(**{field_name: value})],
            filename=f"nonfinite_{field_name}_{value}.csv",
        )

        result = canonicalize_dataset(dataset, _request(dataset), store)

        assert result.canonical_dataset is None
        assert any(field_name in check.check_id for check in result.data_quality_results)


def test_canonical_artifact_preserves_more_than_nine_decimal_places(tmp_path) -> None:
    exact_value = Decimal("10.123456789123")
    dataset, store, _ = _registered_csv(
        tmp_path,
        [_row(line_revenue=str(exact_value), unit_price="1.123456789123")],
        filename="precise.csv",
    )

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert result.canonical_dataset is not None
    parquet_path = store.safe_path(result.canonical_dataset.artifact.path)
    stored = duckdb.sql(f"SELECT line_revenue, unit_price FROM read_parquet('{parquet_path}')").fetchone()
    assert stored == (exact_value, Decimal("1.123456789123"))


def test_unrepresentable_monetary_precision_fails_before_artifact_reference(tmp_path) -> None:
    dataset, store, _ = _registered_csv(
        tmp_path,
        [_row(line_revenue="123456789012345678901234567890123456789")],
        filename="unrepresentable.csv",
    )

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert result.canonical_dataset is None
    assert any("unrepresentable_precision" in check.check_id for check in result.data_quality_results)


def test_dataset_selection_cannot_be_silently_overridden(tmp_path) -> None:
    rows = [_row()]
    dataset, store = _registered_xlsx(tmp_path, rows, selected_sheet="Orders")
    request = _request(dataset)
    request = request.model_copy(update={"selected_sheet": "Other"})

    result = canonicalize_dataset(dataset, request, store)

    assert result.canonical_dataset is None
    assert any(check.check_id == "canonical.selection.conflicting_sheet" for check in result.data_quality_results)


def test_unselected_multisheet_dataset_cannot_use_request_sheet_selection(tmp_path) -> None:
    rows = [_row()]
    dataset, store = _registered_xlsx(tmp_path, rows, selected_sheet=None, extra_sheet="Other")
    request = _request(dataset).model_copy(update={"selected_sheet": "Orders"})

    result = canonicalize_dataset(dataset, request, store)

    assert result.canonical_dataset is None
    assert any(
        check.check_id == "canonical.selection.request_sheet_without_dataset_selection"
        for check in result.data_quality_results
    )


def test_correctly_registered_selected_sheet_proceeds(tmp_path) -> None:
    dataset, store = _registered_xlsx(tmp_path, [_row()], selected_sheet="Orders", extra_sheet="Other")

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert result.canonical_dataset is not None
    assert result.record.selected_sheet == "Orders"


def test_sqlite_dataset_selection_authority_cases(tmp_path) -> None:
    selected_dataset, selected_store = _registered_sqlite(tmp_path, [_row()], selected_table="orders", extra_table=True)
    conflicting = canonicalize_dataset(
        selected_dataset,
        _request(selected_dataset).model_copy(update={"selected_table": "other_orders"}),
        selected_store,
    )
    assert any(check.check_id == "canonical.selection.conflicting_table" for check in conflicting.data_quality_results)

    unselected_dataset, unselected_store = _registered_sqlite(
        tmp_path,
        [_row(order_id="o2")],
        filename="unselected.sqlite",
        selected_table=None,
        extra_table=True,
    )
    ungoverned = canonicalize_dataset(
        unselected_dataset,
        _request(unselected_dataset).model_copy(update={"selected_table": "orders"}),
        unselected_store,
    )
    assert any(
        check.check_id == "canonical.selection.request_table_without_dataset_selection"
        for check in ungoverned.data_quality_results
    )

    ok = canonicalize_dataset(selected_dataset, _request(selected_dataset), selected_store)
    assert ok.canonical_dataset is not None
    assert ok.record.selected_table == "orders"


def test_single_sheet_without_registration_selection_records_unambiguous_sheet(tmp_path) -> None:
    dataset, store = _registered_xlsx(tmp_path, [_row()], selected_sheet=None)

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert result.canonical_dataset is not None
    assert result.record.selected_sheet == "Orders"


def test_csv_and_excel_native_date_only_values_converge(tmp_path) -> None:
    csv_dataset, csv_store, _ = _registered_csv(
        tmp_path,
        [_row(order_date="2026-01-01")],
        filename="date.csv",
    )
    xlsx_dataset, xlsx_store = _registered_xlsx(
        tmp_path,
        [_row(order_date=datetime(2026, 1, 1, 0, 0, 0))],
        selected_sheet="Orders",
    )

    csv_result = canonicalize_dataset(csv_dataset, _request(csv_dataset), csv_store)
    xlsx_result = canonicalize_dataset(xlsx_dataset, _request(xlsx_dataset), xlsx_store)

    assert csv_result.canonical_rows[0].order_date == xlsx_result.canonical_rows[0].order_date


def test_excel_datetime_with_time_component_fails_closed(tmp_path) -> None:
    dataset, store = _registered_xlsx(
        tmp_path,
        [_row(order_date=datetime(2026, 1, 1, 13, 30, 0))],
        selected_sheet="Orders",
    )

    result = canonicalize_dataset(dataset, _request(dataset), store)

    assert result.canonical_dataset is None
    assert any(check.check_id == "canonical.order_date.timestamp_unsupported" for check in result.data_quality_results)


def test_mapped_excel_native_date_column_canonicalizes_to_order_date(tmp_path) -> None:
    dataset, store = _registered_custom_xlsx(
        tmp_path,
        [_mapped_date_row(date_value=datetime(2026, 1, 1, 0, 0, 0))],
        headers=_mapped_date_headers(),
        filename="mapped_date.xlsx",
    )

    result = canonicalize_dataset(dataset, _mapped_date_request(dataset), store)

    assert result.canonical_dataset is not None
    assert result.canonical_rows[0].order_date.isoformat() == "2026-01-01"


def test_mapped_excel_datetime_with_time_component_fails_closed(tmp_path) -> None:
    dataset, store = _registered_custom_xlsx(
        tmp_path,
        [_mapped_date_row(date_value=datetime(2026, 1, 1, 13, 30, 0))],
        headers=_mapped_date_headers(),
        filename="mapped_timestamp.xlsx",
    )

    result = canonicalize_dataset(dataset, _mapped_date_request(dataset), store)

    assert result.canonical_dataset is None
    assert any(check.check_id == "canonical.order_date.timestamp_unsupported" for check in result.data_quality_results)


def test_unmapped_excel_datetime_column_does_not_supply_order_date(tmp_path) -> None:
    headers = ("unmapped_datetime", "DateText", *tuple(header for header in _headers() if header != "order_date"))
    row = _mapped_date_row(source_field="DateText", date_value="01/02/2026")
    row["unmapped_datetime"] = datetime(2026, 1, 1, 0, 0, 0)
    dataset, store = _registered_custom_xlsx(
        tmp_path,
        [row],
        headers=headers,
        filename="unmapped_datetime.xlsx",
    )

    result = canonicalize_dataset(
        dataset,
        _mapped_date_request(dataset, source_field="DateText"),
        store,
    )

    assert result.canonical_dataset is None
    assert any(check.check_id == "canonical.order_date.invalid" for check in result.data_quality_results)


def test_csv_iso_and_mapped_excel_native_dates_converge(tmp_path) -> None:
    csv_dataset, csv_store, _ = _registered_csv(
        tmp_path,
        [_row(order_date="2026-01-01")],
        filename="mapped_converge.csv",
    )
    xlsx_dataset, xlsx_store = _registered_custom_xlsx(
        tmp_path,
        [_mapped_date_row(date_value=datetime(2026, 1, 1, 0, 0, 0))],
        headers=_mapped_date_headers(),
        filename="mapped_converge.xlsx",
    )

    csv_result = canonicalize_dataset(csv_dataset, _request(csv_dataset), csv_store)
    xlsx_result = canonicalize_dataset(xlsx_dataset, _mapped_date_request(xlsx_dataset), xlsx_store)

    assert csv_result.canonical_rows[0].order_date == xlsx_result.canonical_rows[0].order_date


def test_material_canonicalization_configuration_changes_record_identity(tmp_path) -> None:
    dataset, store, _ = _registered_csv(tmp_path, [_row(currency="USD")], filename="identity_config.csv")
    plain = canonicalize_dataset(dataset, _request(dataset), store)
    with_currency_basis = canonicalize_dataset(
        dataset,
        _request(dataset).model_copy(update={"currency_basis_ref": "USD"}),
        store,
    )
    unsupported_date_policy = canonicalize_dataset(
        dataset,
        _request(dataset).model_copy(update={"date_policy_ref": "unsupported"}),
        store,
    )

    assert plain.record.canonicalization_id != with_currency_basis.record.canonicalization_id
    assert unsupported_date_policy.record.canonicalization_id not in {
        plain.record.canonicalization_id,
        with_currency_basis.record.canonicalization_id,
    }
    assert unsupported_date_policy.canonical_dataset is None


def _mapped_date_headers() -> tuple[str, ...]:
    return ("Date", *tuple(header for header in _headers() if header != "order_date"))


def _mapped_date_row(*, source_field: str = "Date", date_value) -> dict[str, object]:
    row = dict(_row())
    row.pop("order_date")
    row[source_field] = date_value
    return row


def _mapped_date_request(dataset, *, source_field: str = "Date") -> CanonicalizationRequest:
    entries = [
        CanonicalFieldMapping(
            canonical_field=canonical_field,
            source_field=source_field if canonical_field == "order_date" else canonical_field,
        )
        for canonical_field in _headers()
    ]
    return CanonicalizationRequest(
        source_dataset_id=dataset.dataset_id,
        mapping=CanonicalMapping(mapping_id=f"map_{source_field}_to_order_date", entries=tuple(entries)),
        eligibility_mode=EligibilityMode.UPSTREAM_ELIGIBLE_ONLY,
    )


def _registered_custom_xlsx(
    tmp_path: Path,
    rows: list[dict[str, object]],
    *,
    headers: tuple[str, ...],
    filename: str,
    selected_sheet: str | None = "Orders",
):
    path = tmp_path / filename
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    workbook.save(path)
    store = ArtifactStore(tmp_path / f"runtime_{path.stem}")
    registry = DatasetRegistry(store)
    return registry.register_source(path, SourceType.EXCEL_XLSX, selected_sheet=selected_sheet), store


def _registered_xlsx(
    tmp_path: Path,
    rows: list[dict[str, str]],
    *,
    numeric: bool = False,
    selected_sheet: str | None = "Orders",
    extra_sheet: str | None = None,
):
    path = tmp_path / f"orders_{selected_sheet or 'unselected'}_{extra_sheet or 'single'}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    headers = _headers()
    sheet.append(headers)
    for row in rows:
        values = []
        for header in headers:
            value = row.get(header, "")
            if numeric and header in {"quantity", "line_revenue", "unit_price"} and value != "":
                value = float(value) if "." in str(value) else int(value)
            values.append(value)
        sheet.append(values)
    if extra_sheet is not None:
        workbook.create_sheet(extra_sheet)
    workbook.save(path)
    store = ArtifactStore(tmp_path / f"runtime_{path.stem}")
    registry = DatasetRegistry(store)
    return registry.register_source(path, SourceType.EXCEL_XLSX, selected_sheet=selected_sheet), store


def _registered_sqlite(
    tmp_path: Path,
    rows: list[dict[str, str]],
    *,
    numeric: bool = False,
    filename: str = "orders.sqlite",
    selected_table: str | None = "orders",
    extra_table: bool = False,
):
    path = tmp_path / filename
    conn = sqlite3.connect(path)
    headers = _headers()
    declared_types = [
        "REAL" if numeric and header in {"quantity", "line_revenue", "unit_price"} else "TEXT"
        for header in headers
    ]
    columns = ", ".join(f'"{header}" {declared_type}' for header, declared_type in zip(headers, declared_types))
    conn.execute("CREATE TABLE orders (" + columns + ")")
    if extra_table:
        conn.execute("CREATE TABLE other_orders (" + columns + ")")
    for row in rows:
        values = []
        for header in headers:
            value = row.get(header, "")
            if numeric and header in {"quantity", "line_revenue", "unit_price"} and value != "":
                value = float(value)
            values.append(value)
        conn.execute(
            "INSERT INTO orders VALUES (" + ", ".join("?" for _ in headers) + ")",
            values,
        )
    conn.commit()
    conn.close()
    store = ArtifactStore(tmp_path / f"runtime_sqlite_{path.stem}")
    registry = DatasetRegistry(store)
    return registry.register_source(path, SourceType.SQLITE, selected_table=selected_table), store
